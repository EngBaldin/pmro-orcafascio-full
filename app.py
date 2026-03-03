import streamlit as st
import pandas as pd
import plotly.express as px
import sqlite3
import io
import re
import pdfplumber
from openpyxl import Workbook
from datetime import datetime, date

st.set_page_config(layout="wide", page_icon="🏗️", page_title="PMRO Enterprise | SEINFRA")

st.markdown("""
<style>
.main {background-color: #f1f5f9;}
.pmro-header {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 50%, #3b82f6 100%);
    padding: 2rem 3rem; border-radius: 16px; color: white; margin-bottom: 1.5rem;
    box-shadow: 0 10px 40px rgba(30,58,138,0.3);
}
.pmro-header h1 {font-size: 2rem; font-weight: 700; margin: 0;}
.pmro-header p {font-size: 0.95rem; margin: 0.3rem 0 0 0; opacity: 0.8;}
.kpi-card {
    background: white; border-radius: 12px; padding: 1.2rem 1.5rem;
    border-left: 4px solid #3b82f6; box-shadow: 0 2px 12px rgba(0,0,0,0.07);
}
.kpi-label {font-size: 0.78rem; color: #64748b; font-weight: 600; text-transform: uppercase;}
.kpi-value {font-size: 1.8rem; font-weight: 700; color: #0f172a; margin-top: 0.2rem;}
.kpi-sub {font-size: 0.78rem; color: #94a3b8;}
.formula-box {
    background: #eff6ff; border: 2px solid #3b82f6; border-radius: 12px;
    padding: 1.5rem; margin: 1rem 0;
}
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def get_db():
    conn = sqlite3.connect("pmro_enterprise.db", check_same_thread=False)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS insumos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT, descricao TEXT, unidade TEXT,
            preco REAL, tabela TEXT, mes_ano TEXT, estado TEXT
        );
        CREATE TABLE IF NOT EXISTS orcamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT UNIQUE, obra TEXT, local TEXT,
            responsavel TEXT, data_orcamento DATE,
            bdi REAL DEFAULT 35.0, subtotal REAL, total_bdi REAL,
            status TEXT DEFAULT 'Rascunho',
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS orcamento_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            orcamento_id INTEGER, codigo TEXT, descricao TEXT,
            unidade TEXT, quantidade REAL, preco_unit REAL, subtotal REAL
        );
        CREATE TABLE IF NOT EXISTS contratos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT UNIQUE, objeto TEXT, empresa TEXT,
            valor REAL, reajuste_indice TEXT, reajuste_base REAL,
            data_base TEXT, data_assinatura TEXT, data_vencimento TEXT,
            status TEXT DEFAULT 'Ativo', pdf_nome TEXT,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    return conn

conn = get_db()

def carregar_insumos():
    return pd.read_sql("SELECT * FROM insumos ORDER BY tabela, codigo", conn)

def carregar_orcamentos():
    return pd.read_sql("SELECT * FROM orcamentos ORDER BY criado_em DESC", conn)

def carregar_contratos():
    return pd.read_sql("SELECT * FROM contratos ORDER BY criado_em DESC", conn)

def importar_excel(file):
    try:
        df = pd.read_excel(file, sheet_name="INSUMOS")
        df.columns = [c.strip().upper() for c in df.columns]
        count = 0
        for _, r in df.iterrows():
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO insumos (codigo,descricao,unidade,preco,tabela,mes_ano,estado) VALUES (?,?,?,?,?,?,?)",
                    (str(r.get("CÓDIGO", r.get("CODIGO",""))).strip(),
                     str(r.get("DESCRIÇÃO", r.get("DESCRICAO",""))).strip(),
                     str(r.get("UNIDADE","")).strip(),
                     float(r.get("PREÇO_UNITARIO", r.get("PRECO_UNITARIO", 0))),
                     str(r.get("TABELA_REFERENCIA","")).strip(),
                     str(r.get("MÊS_ANO", r.get("MES_ANO",""))).strip(),
                     str(r.get("ESTADO","RO")).strip())
                )
                count += 1
            except Exception:
                pass
        conn.commit()
        return True, f"✅ {count} insumos importados com sucesso!"
    except Exception as e:
        return False, f"❌ Erro: {e}"

def gerar_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "INSUMOS"
    ws.append(["CÓDIGO","DESCRIÇÃO","UNIDADE","PREÇO_UNITARIO","TABELA_REFERENCIA","MÊS_ANO","ESTADO"])
    dados = [
        ["INS-001","Cimento Portland CP-II-E-32","kg",0.85,"SINAPI","03/2026","RO"],
        ["INS-002","Areia média lavada","m³",120.00,"SINAPI","03/2026","RO"],
        ["INS-003","Brita nº 1","m³",145.00,"SINAPI","03/2026","RO"],
        ["INS-004","Aço CA-50 ø 10mm","kg",8.50,"SINAPI","03/2026","RO"],
        ["INS-005","Tubo PVC DN 100mm","m",28.40,"SINAPI","03/2026","RO"],
        ["INS-006","Tubo PEAD corrugado DN 200mm","m",156.20,"SINAPI","03/2026","RO"],
        ["INS-007","Tubo PEAD corrugado DN 300mm","m",234.80,"SINAPI","03/2026","RO"],
        ["MOB-001","Pedreiro","h",25.80,"SINAPI","03/2026","RO"],
        ["MOB-002","Servente de obras","h",18.40,"SINAPI","03/2026","RO"],
        ["MOB-003","Carpinteiro","h",28.60,"SINAPI","03/2026","RO"],
        ["EQP-001","Retroescavadeira","h",185.00,"SINAPI","03/2026","RO"],
        ["EQP-002","Motoniveladora Patrol 140H","h",220.00,"SINAPI","03/2026","RO"],
        ["EQP-003","Rolo compactador liso 10t","h",165.00,"SINAPI","03/2026","RO"],
        ["EQP-004","Caminhão basculante 10m³","h",145.00,"SINAPI","03/2026","RO"],
        ["SICRO-001","CBUQ faixa C espessura 4cm","m²",45.67,"SICRO","03/2026","RO"],
        ["SICRO-002","Imprimação betuminosa","m²",6.20,"SICRO","03/2026","RO"],
        ["SICRO-003","Pintura de ligação","m²",4.80,"SICRO","03/2026","RO"],
        ["SICRO-004","Escavação corte motoniveladora","m³",18.90,"SICRO","03/2026","RO"],
        ["PMRO-001","Meio-fio concreto 15x30cm","m",89.50,"BASE_PMRO","03/2026","RO"],
        ["PMRO-002","Sarjeta concreto fck20 L=60cm","m",125.00,"BASE_PMRO","03/2026","RO"],
    ]
    for row in dados:
        ws.append(row)
    ws2 = wb.create_sheet("COMPOSIÇÕES")
    ws2.append(["CÓDIGO_COMP","DESCRIÇÃO_COMPOSIÇÃO","UNIDADE","CÓDIGO_INSUMO","DESCRIÇÃO_INSUMO","COEF","TIPO"])
    ws2.append(["COMP-001","Pavimentação CBUQ 4cm","m²","SICRO-001","CBUQ faixa C",1.0,"MATERIAL"])
    ws2.append(["COMP-001","Pavimentação CBUQ 4cm","m²","SICRO-002","Imprimação",1.0,"MATERIAL"])
    ws2.append(["COMP-001","Pavimentação CBUQ 4cm","m²","EQP-002","Patrol 140H",0.008,"EQUIPAMENTO"])
    ws3 = wb.create_sheet("INSTRUÇÕES")
    ws3.append(["TEMPLATE PMRO Enterprise v6.2 - Preencha INSUMOS e importe no sistema"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# HEADER
st.markdown("""
<div class="pmro-header">
    <h1>🏗️ PMRO Enterprise</h1>
    <p>Sistema de Gestão de Obras Públicas · SEINFRA · Prefeitura Municipal de Porto Velho</p>
</div>
""", unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    st.markdown("### 📂 Navegação")
    pagina = st.radio(
        "Módulos",
        ["📊 Dashboard","💰 Orçamento de Obras","📋 Gestão de Contratos",
         "📅 Planejamento Gantt","📱 Diário de Obras","⚙️ Bases de Dados"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.markdown(f"""
    <div style='text-align:center;color:#64748b;font-size:0.75rem;'>
        <strong>PMRO Enterprise v6.2</strong><br>
        Eng. Guilherme Ritter Baldin<br>
        SEINFRA · Porto Velho · RO<br>
        {datetime.now().strftime("%d/%m/%Y %H:%M")}
    </div>
    """, unsafe_allow_html=True)

# ═══════════════════════════════
# DASHBOARD
# ═══════════════════════════════
if pagina == "📊 Dashboard":
    df_orc = carregar_orcamentos()
    df_cont = carregar_contratos()
    df_ins = carregar_insumos()
    total_orc = df_orc["total_bdi"].sum() if not df_orc.empty else 0
    total_cont = df_cont["valor"].sum() if not df_cont.empty else 0

    col1, col2, col3, col4 = st.columns(4)
    col1.markdown(f'<div class="kpi-card"><div class="kpi-label">📋 Orçamentos</div><div class="kpi-value">{len(df_orc)}</div><div class="kpi-sub">Elaborados</div></div>', unsafe_allow_html=True)
    col2.markdown(f'<div class="kpi-card"><div class="kpi-label">💰 Valor Orçado</div><div class="kpi-value">R$ {total_orc:,.0f}</div><div class="kpi-sub">Com BDI</div></div>', unsafe_allow_html=True)
    col3.markdown(f'<div class="kpi-card"><div class="kpi-label">📄 Contratos</div><div class="kpi-value">{len(df_cont)}</div><div class="kpi-sub">R$ {total_cont:,.0f}</div></div>', unsafe_allow_html=True)
    col4.markdown(f'<div class="kpi-card"><div class="kpi-label">🗂️ Insumos Base</div><div class="kpi-value">{len(df_ins)}</div><div class="kpi-sub">SINAPI+SICRO+PMRO</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if not df_orc.empty:
        fig = px.bar(df_orc.head(8), x="numero", y="total_bdi",
                     title="Orçamentos por Valor (R$)", color="status",
                     labels={"total_bdi":"Valor R$","numero":"Orçamento"})
        fig.update_layout(height=350, plot_bgcolor="white", paper_bgcolor="white")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("💡 Importe insumos em **⚙️ Bases de Dados** e crie seu primeiro orçamento!")

# ═══════════════════════════════
# ORÇAMENTO
# ═══════════════════════════════
elif pagina == "💰 Orçamento de Obras":
    tabs = st.tabs(["➕ Novo Orçamento","📋 Orçamentos Salvos"])

    with tabs[0]:
        st.subheader("Novo Orçamento PMRO")
        col1, col2, col3 = st.columns(3)
        num_orc = col1.text_input("Nº Orçamento*", placeholder="ORC-2026-001")
        nome_obra = col2.text_input("Nome da Obra*", placeholder="Pavimentação Av. Guaporé")
        local_obra = col3.text_input("Localização", placeholder="Bairro 3 Marias")
        resp = col1.text_input("Responsável", value="Eng. Guilherme Ritter Baldin")
        bdi = col2.number_input("BDI (%)", value=35.0, min_value=0.0, max_value=100.0)
        data_orc = col3.date_input("Data", value=date.today())

        st.markdown("---")
        st.markdown("#### 🔍 Pesquisar Insumo na Base")
        df_ins = carregar_insumos()
        if df_ins.empty:
            st.warning("⚠️ Base vazia. Importe insumos em ⚙️ Bases de Dados primeiro.")
        else:
            col_b1, col_b2 = st.columns([3,1])
            busca = col_b1.text_input("Buscar (nome ou código):", placeholder="ex: PEAD, asfalto, pedreiro")
            tab_f = col_b2.selectbox("Tabela:", ["TODAS"] + sorted(df_ins["tabela"].unique().tolist()))
            df_f = df_ins.copy()
            if busca:
                df_f = df_f[df_f["descricao"].str.contains(busca, case=False, na=False) | df_f["codigo"].str.contains(busca, case=False, na=False)]
            if tab_f != "TODAS":
                df_f = df_f[df_f["tabela"] == tab_f]
            if not df_f.empty:
                st.dataframe(df_f[["codigo","descricao","unidade","preco","tabela"]].head(15),
                    use_container_width=True, hide_index=True,
                    column_config={"preco": st.column_config.NumberColumn("Preço Unit.", format="R$ %.2f")})

        st.markdown("#### ➕ Adicionar Item")
        if "itens" not in st.session_state:
            st.session_state.itens = []

        col1i, col2i, col3i, col4i, col5i = st.columns([2,4,1,2,2])
        cod_i = col1i.text_input("Código", key="ci")
        desc_i = col2i.text_input("Descrição", key="di")
        und_i = col3i.text_input("Und", key="ui")
        qtd_i = col4i.number_input("Qtde", min_value=0.0, step=1.0, key="qi")
        preco_i = col5i.number_input("Preço Unit.", min_value=0.0, step=0.01, key="pi")

        if not df_ins.empty and cod_i:
            match = df_ins[df_ins["codigo"].str.strip() == cod_i.strip()]
            if not match.empty:
                st.success(f"✅ {match.iloc[0]['descricao']} | {match.iloc[0]['unidade']} | R$ {match.iloc[0]['preco']:.2f}")

        if st.button("➕ Adicionar Item", type="primary"):
            if cod_i and desc_i and qtd_i > 0 and preco_i > 0:
                st.session_state.itens.append({
                    "codigo": cod_i, "descricao": desc_i, "unidade": und_i,
                    "quantidade": qtd_i, "preco_unit": preco_i, "subtotal": qtd_i * preco_i
                })
                st.rerun()

        if st.session_state.itens:
            df_itens = pd.DataFrame(st.session_state.itens)
            subtotal = df_itens["subtotal"].sum()
            total_final = subtotal * (1 + bdi / 100)
            st.dataframe(df_itens, use_container_width=True, hide_index=True,
                column_config={
                    "subtotal": st.column_config.NumberColumn("Subtotal R$", format="R$ %.2f"),
                    "preco_unit": st.column_config.NumberColumn("Unit. R$", format="R$ %.2f")
                })
            col_s, col_b2, col_t = st.columns(3)
            col_s.metric("Subtotal", f"R$ {subtotal:,.2f}")
            col_b2.metric(f"BDI {bdi}%", f"R$ {subtotal*(bdi/100):,.2f}")
            col_t.metric("TOTAL GERAL", f"R$ {total_final:,.2f}")

            col_sv, col_ex, col_cl = st.columns(3)
            if col_sv.button("💾 Salvar Orçamento", type="primary"):
                try:
                    conn.execute(
                        "INSERT INTO orcamentos (numero,obra,local,responsavel,data_orcamento,bdi,subtotal,total_bdi) VALUES (?,?,?,?,?,?,?,?)",
                        (num_orc, nome_obra, local_obra, resp, str(data_orc), bdi, subtotal, total_final)
                    )
                    orc_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    for item in st.session_state.itens:
                        conn.execute(
                            "INSERT INTO orcamento_itens (orcamento_id,codigo,descricao,unidade,quantidade,preco_unit,subtotal) VALUES (?,?,?,?,?,?,?)",
                            (orc_id, item["codigo"], item["descricao"], item["unidade"], item["quantidade"], item["preco_unit"], item["subtotal"])
                        )
                    conn.commit()
                    st.balloons()
                    st.success(f"✅ Orçamento {num_orc} salvo! Total: R$ {total_final:,.2f}")
                    st.session_state.itens = []
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ {e}")

            if col_ex.button("📥 Exportar Excel"):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_itens.to_excel(writer, sheet_name="Planilha Orçamentária", index=False)
                    pd.DataFrame({
                        "Resumo": ["Subtotal", f"BDI {bdi}%", "Total Geral"],
                        "R$": [subtotal, subtotal*(bdi/100), total_final]
                    }).to_excel(writer, sheet_name="Resumo", index=False)
                st.download_button("⬇️ Baixar Excel PMRO", output.getvalue(),
                    f"PMRO_{num_orc}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if col_cl.button("🗑️ Limpar"):
                st.session_state.itens = []
                st.rerun()

    with tabs[1]:
        df_orc = carregar_orcamentos()
        if not df_orc.empty:
            st.dataframe(df_orc, use_container_width=True, hide_index=True,
                column_config={"total_bdi": st.column_config.NumberColumn("Total R$", format="R$ %.2f")})
            st.download_button("📥 CSV", df_orc.to_csv(index=False).encode(), "PMRO_Orcamentos.csv")
        else:
            st.info("Nenhum orçamento salvo.")

# ═══════════════════════════════
# CONTRATOS
# ═══════════════════════════════
elif pagina == "📋 Gestão de Contratos":
    tabs = st.tabs(["➕ Novo Contrato","🔄 Cálculo de Reajuste","📋 Todos Contratos"])

    with tabs[0]:
        st.subheader("Cadastro de Contrato")
        uploaded = st.file_uploader("📄 Upload PDF do Contrato (IA extrai dados)", type="pdf")
        numero_c, valor_c, empresa_c, objeto_c = "", 0.0, "", ""

        if uploaded:
            try:
                texto = ""
                with pdfplumber.open(uploaded) as pdf:
                    for p in pdf.pages:
                        texto += p.extract_text() or ""
                n = re.search(r"(\d{3}/PGM/\d{4})", texto)
                v = re.search(r"valor.*?R\$\s*([\d.,]+)", texto, re.I)
                numero_c = n.group(1) if n else ""
                valor_c = float(v.group(1).replace(".", "").replace(",", ".")) if v else 0.0
                st.success("✅ PDF lido com sucesso!")
                with st.expander("🔍 Texto extraído do PDF"):
                    st.text_area("Conteúdo extraído:", texto[:2000], height=150)
            except Exception as e:
                st.warning(f"Leitura parcial: {e}")

        col1, col2 = st.columns(2)
        numero = col1.text_input("Nº Contrato*", value=numero_c, placeholder="047/PGM/2025")
        empresa = col2.text_input("Empresa Contratada*", value=empresa_c)
        objeto = st.text_area("Objeto do Contrato", value=objeto_c, height=80)

        col1, col2, col3 = st.columns(3)
        valor = col1.number_input("Valor Contratual R$*", value=float(valor_c), min_value=0.0, step=1000.0)
        dt_ass = col2.text_input("Data Assinatura (dd/mm/aaaa)", value=date.today().strftime("%d/%m/%Y"))
        dt_venc = col3.text_input("Data Vencimento (dd/mm/aaaa)", value="")

        st.markdown("#### 📅 Data-Base e Índice de Reajuste")
        st.info("⚠️ A data-base é vinculada à data do **orçamento estimado** da Administração (Lei 14.133/2021, Art. 92)")
        col1, col2, col3 = st.columns(3)
        dt_base = col1.text_input("📅 Data-Base do Orçamento Estimado*", value=date.today().strftime("%d/%m/%Y"))
        indice = col2.selectbox("Índice de Reajuste", ["SINAPI-RO","SICRO-DNIT","SINAPI-RO + SICRO-DNIT","INCC-DI","IPCA","IGP-M"])
        reajuste_base_val = col3.number_input("Índice Io (na data-base)*", value=100.0, step=0.0001, format="%.4f",
                                               help="Valor do índice SINAPI/SICRO na data do orçamento estimado")

        if st.button("💾 Salvar Contrato", type="primary"):
            if not numero or valor <= 0:
                st.error("❌ Preencha Nº Contrato e Valor obrigatoriamente.")
            else:
                try:
                    conn.execute(
                        "INSERT INTO contratos (numero,objeto,empresa,valor,reajuste_indice,reajuste_base,data_base,data_assinatura,data_vencimento,pdf_nome) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (numero, objeto, empresa, valor, indice, reajuste_base_val, dt_base, dt_ass, dt_venc, uploaded.name if uploaded else "")
                    )
                    conn.commit()
                    st.balloons()
                    st.success(f"✅ Contrato {numero} salvo! Valor: R$ {valor:,.2f}")
                except Exception as e:
                    st.error(f"❌ {e}")

    with tabs[1]:
        st.subheader("🔄 Cálculo de Reajuste Contratual")
        st.markdown("""
        <div class="formula-box">
            <h4 style="color:#1e3a8a;margin:0 0 0.5rem 0;">📐 Fórmula Oficial — Lei 14.133/2021</h4>
            <h3 style="text-align:center;color:#0f172a;">R = ((Ii – Io) / Io) × V</h3>
            <p style="margin:0.5rem 0 0 0;color:#475569;">
            <strong>R</strong> = Valor do reajustamento &nbsp;|&nbsp;
            <strong>Ii</strong> = Índice do mês de reajustamento &nbsp;|&nbsp;
            <strong>Io</strong> = Índice da data-base &nbsp;|&nbsp;
            <strong>V</strong> = Valor remanescente a reajustar
            </p>
        </div>
        """, unsafe_allow_html=True)

        df_cont = carregar_contratos()
        if not df_cont.empty:
            sel = st.selectbox("Selecione o Contrato:", df_cont["numero"].tolist())
            c = df_cont[df_cont["numero"] == sel].iloc[0]

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Valor Contratual", f"R$ {float(c['valor']):,.2f}")
            col2.metric("Índice Io (base)", f"{float(c['reajuste_base']):.4f}")
            col3.metric("Índice Referência", str(c["reajuste_indice"]))
            col4.metric("Data-Base", str(c["data_base"]))

            st.markdown("---")
            st.markdown("#### Parâmetros do Reajuste")
            col1, col2 = st.columns(2)
            V = col1.number_input(
                "💰 V — Valor remanescente a reajustar (R$):",
                value=float(c["valor"]), step=1000.0,
                help="Somente obrigações iniciadas e concluídas APÓS a anualidade"
            )
            Ii = col2.number_input(
                "📈 Ii — Índice atual (mês do reajustamento):",
                value=float(c["reajuste_base"]), step=0.0001, format="%.4f",
                help="Consulte o índice no IBGE (SINAPI) ou DNIT (SICRO)"
            )
            Io = float(c["reajuste_base"])

            if Ii > 0 and Io > 0 and V > 0:
                R = ((Ii - Io) / Io) * V
                variacao_pct = ((Ii - Io) / Io) * 100
                valor_total = V + R

                st.markdown("#### 📊 Resultado")
                col1, col2, col3 = st.columns(3)
                col1.metric("R — Reajustamento", f"R$ {R:,.2f}", f"{variacao_pct:+.4f}%")
                col2.metric("Variação do Índice", f"{variacao_pct:.4f}%")
                col3.metric("Novo Valor Total", f"R$ {valor_total:,.2f}")

                st.markdown("#### 📋 Memória de Cálculo Oficial")
                memoria = f"""MEMÓRIA DE CÁLCULO — REAJUSTE CONTRATUAL
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Contrato Nº:     {c['numero']}
Empresa:         {c['empresa']}
Índice Adotado:  {c['reajuste_indice']}
Data-Base (Io):  {c['data_base']}

FÓRMULA:  R = ((Ii – Io) / Io) × V

Ii  (Índice mês reajuste):    {Ii:.4f}
Io  (Índice data-base):       {Io:.4f}
V   (Valor remanescente):     R$ {V:,.2f}

CÁLCULO:
R = (({Ii:.4f} – {Io:.4f}) / {Io:.4f}) × R$ {V:,.2f}
R = ({Ii - Io:.4f} / {Io:.4f}) × R$ {V:,.2f}
R = {(Ii - Io)/Io:.6f} × R$ {V:,.2f}
R = R$ {R:,.2f}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
VALOR DO REAJUSTAMENTO (R):    R$ {R:,.2f}
VALOR TOTAL REAJUSTADO (V+R):  R$ {valor_total:,.2f}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Elaborado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}
Responsável:  Eng. Guilherme Ritter Baldin — SEINFRA/PMRO"""
                st.code(memoria)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    pd.DataFrame({
                        "Parâmetro": ["Contrato","Empresa","Índice","Data-Base","Ii","Io","V","R","V+R","Data Cálculo"],
                        "Valor": [c['numero'], c['empresa'], c['reajuste_indice'], str(c['data_base']),
                                  f"{Ii:.4f}", f"{Io:.4f}", f"R$ {V:,.2f}", f"R$ {R:,.2f}",
                                  f"R$ {valor_total:,.2f}", datetime.now().strftime("%d/%m/%Y")]
                    }).to_excel(writer, sheet_name="Memória de Cálculo", index=False)
                st.download_button(
                    "📥 Exportar Memória de Cálculo Excel",
                    output.getvalue(),
                    f"Reajuste_{str(c['numero']).replace('/','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Cadastre contratos na aba anterior.")

    with tabs[2]:
        df_cont = carregar_contratos()
        if not df_cont.empty:
            st.dataframe(df_cont, use_container_width=True, hide_index=True,
                column_config={"valor": st.column_config.NumberColumn("Valor R$", format="R$ %.2f")})
            st.download_button("📥 CSV", df_cont.to_csv(index=False).encode(), "PMRO_Contratos.csv")
        else:
            st.info("Nenhum contrato cadastrado.")

# ═══════════════════════════════
# GANTT
# ═══════════════════════════════
elif pagina == "📅 Planejamento Gantt":
    st.subheader("📅 Planejamento Gantt")
    st.info("🚧 Em desenvolvimento — próxima versão: cronograma físico-financeiro interativo com curva ABC 112/2009.")

# ═══════════════════════════════
# DIÁRIO
# ═══════════════════════════════
elif pagina == "📱 Diário de Obras":
    st.subheader("📱 Diário de Obras")
    st.info("🚧 Em desenvolvimento — próxima versão: registro diário, equipe, chuvas, relatório mensal automático.")

# ═══════════════════════════════
# BASES DE DADOS
# ═══════════════════════════════
elif pagina == "⚙️ Bases de Dados":
    tabs = st.tabs(["📥 Importar Tabela","🔍 Ver Base","➕ Insumo Manual"])

    with tabs[0]:
        st.subheader("Importar Tabela de Referência")
        st.info("📌 Baixe o template, preencha com SINAPI/SICRO/ORSE/BASE_PMRO e importe.")
        st.download_button(
            "⬇️ Baixar Template Excel PMRO",
            gerar_template(),
            "TEMPLATE_PMRO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        st.markdown("---")
        upload_excel = st.file_uploader("📊 Upload Tabela (.xlsx)", type=["xlsx"])
        if upload_excel:
            ok, msg = importar_excel(upload_excel)
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    with tabs[1]:
        df_ins = carregar_insumos()
        if not df_ins.empty:
            col1, col2 = st.columns(2)
            busca = col1.text_input("🔍 Buscar:")
            tab_f = col2.selectbox("Tabela:", ["TODAS"] + sorted(df_ins["tabela"].unique().tolist()))
            df_s = df_ins.copy()
            if busca:
                df_s = df_s[df_s["descricao"].str.contains(busca, case=False, na=False)]
            if tab_f != "TODAS":
                df_s = df_s[df_s["tabela"] == tab_f]
            st.metric("Insumos encontrados", len(df_s))
            st.dataframe(df_s, use_container_width=True, hide_index=True,
                column_config={"preco": st.column_config.NumberColumn("Preço R$", format="R$ %.2f")})
        else:
            st.info("Base vazia. Importe uma planilha.")

    with tabs[2]:
        st.subheader("Adicionar Insumo Manual")
        col1, col2, col3 = st.columns(3)
        m_cod = col1.text_input("Código")
        m_desc = col2.text_input("Descrição")
        m_und = col3.text_input("Unidade")
        col1, col2, col3 = st.columns(3)
        m_preco = col1.number_input("Preço R$", min_value=0.0, step=0.01)
        m_tab = col2.selectbox("Tabela", ["SINAPI","SICRO","ORSE","SEINFRA","BASE_PMRO"])
        m_mes = col3.text_input("Mês/Ano", value=datetime.now().strftime("%m/%Y"))
        if st.button("➕ Adicionar", type="primary"):
            if m_cod and m_desc and m_preco > 0:
                conn.execute(
                    "INSERT INTO insumos (codigo,descricao,unidade,preco,tabela,mes_ano,estado) VALUES (?,?,?,?,?,?,?)",
                    (m_cod, m_desc, m_und, m_preco, m_tab, m_mes, "RO")
                )
                conn.commit()
                st.success(f"✅ {m_cod} adicionado!")
                st.rerun()

st.markdown("---")
st.markdown(f"""
<div style='text-align:center;color:#94a3b8;font-size:0.8rem;padding:1rem;'>
    🏗️ <strong>PMRO Enterprise v6.2</strong> · SEINFRA · Prefeitura Municipal de Porto Velho<br>
    © {datetime.now().year} Eng. Guilherme Ritter Baldin · Todos os direitos reservados
</div>
""", unsafe_allow_html=True)
